from pathlib import Path
import logging
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ===== パス設定 =====
BASE_DIR = Path(__file__).parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
LOG_DIR = BASE_DIR / "logs"

INPUT_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)

# 入出力（フォルダ構成に合わせる）
CSV_PATH = INPUT_DIR / "sales.csv"
OUT_XLSX = OUTPUT_DIR / "report.xlsx"


# ===== ログ設定 =====
log_file = LOG_DIR / "run.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler()
    ]
)


def ensure_sample_csv(path: Path) -> None:
    """sales.csv が無ければサンプルを作る"""
    if path.exists():
        return

    df = pd.DataFrame(
        [
            {"date": "2026-01-01", "product": "Apple",  "qty": 2, "unit_price": 120},
            {"date": "2026-01-01", "product": "Orange", "qty": 5, "unit_price": 80},
            {"date": "2026-01-02", "product": "Apple",  "qty": 1, "unit_price": 120},
            {"date": "2026-01-03", "product": "Banana", "qty": 10, "unit_price": 50},
            {"date": "2026-01-03", "product": "Orange", "qty": 3, "unit_price": 80},
            {"date": "2026-02-01", "product": "Apple",  "qty": 4, "unit_price": 120},
            {"date": "2026-02-02", "product": "Banana", "qty": 6, "unit_price": 50},
        ]
    )
    df.to_csv(path, index=False, encoding="utf-8")
    logging.info("サンプルCSVを作成しました: %s", path)


def make_summary_tables(csv_path: Path):
    df = pd.read_csv(csv_path)
    df["date"] = pd.to_datetime(df["date"])
    df["month"] = df["date"].dt.to_period("M").astype(str)
    df["sales"] = df["qty"] * df["unit_price"]

    monthly = (
        df.groupby("month", as_index=False)["sales"]
        .sum()
        .sort_values("month")
        .rename(columns={"sales": "monthly_sales"})
    )

    by_product = (
        df.groupby("product", as_index=False)["sales"]
        .sum()
        .sort_values("sales", ascending=False)
        .rename(columns={"sales": "product_sales"})
    )

    total = int(df["sales"].sum())
    return monthly, by_product, total


def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def style_sheet(ws, title: str):
    ws.insert_rows(1)
    ws["A1"] = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    ws.freeze_panes = "A3"
    autosize_columns(ws)


def apply_number_format(ws, col_name: str, fmt: str):
    headers = [c.value for c in ws[2]]
    if col_name not in headers:
        return
    col_idx = headers.index(col_name) + 1

    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=col_idx).number_format = fmt


def export_excel(monthly, by_product, total, out_xlsx: Path):
    # pandasで中身作成
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        monthly.to_excel(writer, sheet_name="Monthly", index=False)
        by_product.to_excel(writer, sheet_name="ByProduct", index=False)

    # openpyxlで見た目を整える
    wb = load_workbook(out_xlsx)

    ws_m = wb["Monthly"]
    style_sheet(ws_m, "月別 売上レポート")
    apply_number_format(ws_m, "monthly_sales", "#,##0")

    ws_p = wb["ByProduct"]
    style_sheet(ws_p, "商品別 売上レポート")
    apply_number_format(ws_p, "product_sales", "#,##0")

    ws_s = wb.create_sheet("Summary", 0)
    ws_s["A1"] = "サマリー"
    ws_s["A1"].font = Font(bold=True, size=14)
    ws_s["A3"] = "総売上"
    ws_s["B3"] = total
    ws_s["B3"].number_format = "#,##0"
    ws_s["A3"].font = Font(bold=True)
    ws_s["B3"].alignment = Alignment(horizontal="right")

    autosize_columns(ws_s)

    wb.save(out_xlsx)


def main():
    logging.info("===== レポート作成 開始 =====")

    try:
        ensure_sample_csv(CSV_PATH)

        logging.info("処理対象CSV：%s", CSV_PATH.name)
        monthly, by_product, total = make_summary_tables(CSV_PATH)

        export_excel(monthly, by_product, total, OUT_XLSX)
        logging.info("Excel出力完了：%s", OUT_XLSX)

        logging.info("正常終了しました")

    except Exception:
        logging.exception("異常終了しました")

    finally:
        logging.info("===== 処理終了 =====")


if __name__ == "__main__":
    main()
